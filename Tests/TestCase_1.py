import openpyxl
import csv
import nose

from selenium import webdriver as WD

class TestCase_1(object):
    """To verify User can Perform Sorting by """
    def __init__(self, path_Xl_Sheet, part_Of_Objective):
        super(TestCase_1, self).__init__()
        self.path_Xl_Sheet = path_Xl_Sheet
        self.TestCase, self.TestSteps = self.openXlsDoc(part_Of_Objective)

        # self.formated_Test_Cases = self.create_Test_Steps()
        # print self.__dict__
        # print self.__doc__

    def openXlsDoc(self, part_Of_Objective):
        WB = openpyxl.load_workbook(self.path_Xl_Sheet, use_iterators = True)

        Events_Sheet = WB.get_sheet_by_name(name = "Events")
        cnt = 0

        Copy_Sheet = []

        for row in Events_Sheet.iter_rows():
            Copy_Sheet.append(row)
            # val = []
            # for col in Events_Sheet.iter_cols():
            #     val.append(Events_Sheet.cell(row, col).value)
            # Copy_Sheet.append(val)

        for i in Copy_Sheet:
            # print i[1].value
            if i[1].value == (TestCase_1.__doc__+str(part_Of_Objective)):
                self.__doc__ = i[1].value
                return i[0].value, i[3].value
            cnt += 1

        # for row in range(Events_Sheet.nrows):

        #     print cnt,(TestCase_1.__doc__+"Name on Events Page"), Events_Sheet.cell(row, 1)
        #     if Events_Sheet.cell(row, 1).value == (TestCase_1.__doc__+"Name on Events Page"):
        #         self.__doc__ = Events_Sheet.cell(row, 1).value
        #         return Events_Sheet.cell(row, 1).value, Events_Sheet.cell(row, 3).value
        #     cnt += 1
        return "a", "b"

    def create_Test_Steps(self):
        TestStep = list()
        while self.TestSteps != "":
            b, c, self.TestSteps = self.TestSteps.partition(". ")
            b = b.rstrip(" 1234567890 \n")

            #TODO: remove "(" ")" and "."
            # b = b.remove("( ) .")

            if not b.isalpha() and not b != "":
                # TestStep = TestStep + "\n"
                pass
            else:
                TestStep.append( b.replace(" ", "_") )
        return TestStep
        # print "b:", b+"_", "\na", a.TestSteps, "\nc", c, "\nTestStep", TestStep


# a = TestCase_1("../SocialTables_tBlocks Test cases sheet V1.0.xlsx", "Name on Events Page")
# print list(a.TestSteps.format( "Open" ) )

# print a.formated_Test_Cases



# print TestStep


